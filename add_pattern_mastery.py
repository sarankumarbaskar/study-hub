import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

fp = "/Users/sbaskar/Personal/Interview-Preparation-Tracker.xlsx"
wb = openpyxl.load_workbook(fp)

C_BG="0A0E17"; C_DARK="111827"; C_CARD="1A2234"; C_BORDER="2D3748"
C_TEXT="E2E8F0"; C_MUTED="94A3B8"; C_ACCENT="60A5FA"; C_GREEN="34D399"
C_YELLOW="FBBF24"; C_ORANGE="FB923C"; C_RED="F87171"; C_PINK="F472B6"
C_PURPLE="A78BFA"; C_WHITE="FFFFFF"; C_CYAN="22D3EE"

FILL_H=PatternFill("solid",fgColor=C_ACCENT); FILL_D=PatternFill("solid",fgColor=C_DARK)
FILL_C=PatternFill("solid",fgColor=C_CARD); FILL_BG=PatternFill("solid",fgColor=C_BG)
FILL_SEC=PatternFill("solid",fgColor="1E3A5F"); FILL_GRN=PatternFill("solid",fgColor="1A3A2A")
FILL_YEL=PatternFill("solid",fgColor="3A3520"); FILL_RED=PatternFill("solid",fgColor="3A1A1A")
FILL_READY=PatternFill("solid",fgColor="0D3320"); FILL_WARN=PatternFill("solid",fgColor="3A2A10")

FH=Font("Calibri",bold=True,size=11,color=C_WHITE)
FT=Font("Calibri",bold=True,size=16,color=C_ACCENT)
FS=Font("Calibri",bold=True,size=12,color=C_TEXT)
FS2=Font("Calibri",bold=True,size=11,color=C_ACCENT)
FB=Font("Calibri",size=11,color=C_TEXT)
FM=Font("Calibri",size=10,color=C_MUTED)
FG=Font("Calibri",bold=True,size=11,color=C_GREEN)
FY=Font("Calibri",bold=True,size=11,color=C_YELLOW)
FR=Font("Calibri",bold=True,size=11,color=C_RED)
FP=Font("Calibri",bold=True,size=11,color=C_PURPLE)
FC=Font("Calibri",bold=True,size=11,color=C_CYAN)
F_CHK=Font("Calibri",size=11,color=C_MUTED)
F_SEC_T=Font("Calibri",bold=True,size=13,color=C_ACCENT)
F_LEGEND=Font("Calibri",size=10,color=C_TEXT)
F_LEGEND_H=Font("Calibri",bold=True,size=10,color=C_ACCENT)
F_STAT_L=Font("Calibri",size=11,color=C_MUTED)
F_STAT_V=Font("Calibri",bold=True,size=14,color=C_GREEN)
F_STAT_V2=Font("Calibri",bold=True,size=14,color=C_YELLOW)

BD=Border(left=Side("thin",color=C_BORDER),right=Side("thin",color=C_BORDER),top=Side("thin",color=C_BORDER),bottom=Side("thin",color=C_BORDER))
AC=Alignment(horizontal="center",vertical="center",wrap_text=True)
AL=Alignment(horizontal="left",vertical="center",wrap_text=True)

def hdr(ws,r,n):
    for c in range(1,n+1):
        cl=ws.cell(r,c); cl.font=FH; cl.fill=FILL_H; cl.alignment=AC; cl.border=BD

def body(ws,r,n,alt=False):
    f=FILL_C if alt else FILL_D
    for c in range(1,n+1):
        cl=ws.cell(r,c); cl.font=FB; cl.fill=f; cl.alignment=AC; cl.border=BD

def sec_title(ws,r,ncols,text):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ncols)
    cl=ws.cell(r,1); cl.value=text; cl.font=F_SEC_T; cl.fill=FILL_SEC; cl.alignment=AL; cl.border=BD

def dv(ws,col,r1,r2,opts):
    v=DataValidation(type="list",formula1=f'"{",".join(opts)}"',allow_blank=True)
    ws.add_data_validation(v); v.add(f"{col}{r1}:{col}{r2}")

def widths(ws,w_list):
    for i,w in enumerate(w_list,1): ws.column_dimensions[get_column_letter(i)].width=w

# ═══════════════════════════════════════════════════════════════════════════
# PATTERNS DATA
# ═══════════════════════════════════════════════════════════════════════════

PATTERNS = [
    "Arrays & Hashing", "Prefix Sum", "Two Pointers", "Sliding Window",
    "Binary Search", "Bit Manipulation", "Stack", "Monotonic Stack",
    "Queue", "Heap / Priority Queue", "Linked List", "Fast & Slow Pointer",
    "Trees (DFS)", "Trees (BFS)", "Binary Search Tree",
    "Graph DFS", "Graph BFS", "Topological Sort", "Union Find (Disjoint Set)", "Trie",
    "Backtracking", "Greedy", "Dynamic Programming",
    "Interval Problems", "Matrix Problems", "Design Problems",
]

CHECKLISTS = {
    "Arrays & Hashing": ["HashMap for O(1) lookup","Frequency counting","Prefix/suffix product","Sorting + scan","Bucket sort","Two-pass techniques","In-place modification","Duplicate detection"],
    "Prefix Sum": ["Basic prefix sum","Range sum queries","Subarray sum equals K","2D prefix sum","Prefix XOR","Difference array","Running average"],
    "Two Pointers": ["Opposite-end pointers","Same-direction pointers","Dutch National Flag","Three pointers (3Sum)","Expand from center","Merge two sorted","Partition in-place"],
    "Sliding Window": ["Fixed-size window","Variable-size window","Character frequency window","Distinct elements window","Window shrinking condition","Monotonic queue in window","Prefix sum + window","Binary search + window"],
    "Binary Search": ["Classic binary search","Lower bound (leftmost)","Upper bound (rightmost)","Rotated sorted arrays","Binary search on answer","Peak/valley problems","Infinite search space","Matrix binary search"],
    "Bit Manipulation": ["XOR for uniqueness","Brian Kernighan's trick","Bit counting","Carry-free addition","Power of 2 checks","Bit masking","Subset generation via bits"],
    "Stack": ["Balanced parentheses","Expression evaluation","Nested structure parsing","Undo/redo simulation","Min/max stack","Function call simulation"],
    "Monotonic Stack": ["Next greater element","Next smaller element","Largest rectangle","Stock span","Temperature problems","Sum of subarray min/max","Contribution technique"],
    "Queue": ["BFS level-order","Circular queue","Priority scheduling","Sliding window via deque","Task processing order"],
    "Heap / Priority Queue": ["Kth largest/smallest","Top K elements","Two-heap median","Merge K sorted","Event-driven simulation","Greedy + heap combo","Lazy deletion"],
    "Linked List": ["Iterative reversal","Recursive reversal","Dummy head technique","Two-pointer (gap)","Deep copy with random","Merge sorted lists","K-group reversal"],
    "Fast & Slow Pointer": ["Cycle detection (Floyd's)","Cycle start finding","Middle of list","Happy number","Duplicate in array"],
    "Trees (DFS)": ["Preorder traversal","Inorder traversal","Postorder traversal","Path sum problems","Tree diameter","Maximum path sum","Subtree problems","Tree serialization"],
    "Trees (BFS)": ["Level-order traversal","Right/left side view","Zigzag traversal","Minimum depth","Connect next pointers","Level averages"],
    "Binary Search Tree": ["BST validation","Inorder = sorted property","Kth smallest/largest","LCA in BST","BST insert/delete","Balanced BST construction"],
    "Graph DFS": ["Connected components","Island counting","Flood fill","Cycle detection (directed)","Cycle detection (undirected)","Path existence","All paths source to target","Articulation points"],
    "Graph BFS": ["Shortest path unweighted","Multi-source BFS","Rotting oranges pattern","Word ladder pattern","Level-by-level processing","0-1 BFS"],
    "Topological Sort": ["Kahn's algorithm (BFS)","DFS-based topo sort","Cycle detection in DAG","Course scheduling","Build order","Alien dictionary"],
    "Union Find (Disjoint Set)": ["Path compression","Union by rank","Connected components count","Redundant edge detection","Accounts merge pattern","Dynamic connectivity","MST (Kruskal's)"],
    "Trie": ["Insert/search/prefix","Wildcard matching with DFS","Word search in grid","Autocomplete system","Longest common prefix","Count words with prefix"],
    "Backtracking": ["Subsets generation","Permutations","Combinations","Constraint satisfaction","Pruning strategies","State restoration","Grid path exploration","N-Queens pattern"],
    "Greedy": ["Activity selection","Jump game pattern","Interval scheduling","Huffman-style greedy","Exchange argument proof","Kadane's algorithm","Task scheduling"],
    "Dynamic Programming": ["1D linear DP","2D grid DP","Knapsack (0/1 and unbounded)","String DP (LCS/edit distance)","Interval DP","State machine DP (stocks)","Bitmask DP","Digit DP","Tree DP"],
    "Interval Problems": ["Sort by start time","Sort by end time","Merge overlapping","Insert interval","Non-overlapping selection","Line sweep","Meeting rooms pattern"],
    "Matrix Problems": ["Rotation (90/180/270)","Spiral traversal","Set matrix zeroes","Search in 2D matrix","Diagonal traversal","Layer-by-layer processing"],
    "Design Problems": ["LRU/LFU cache","HashMap from scratch","Trie-based design","Random O(1) insert/delete","Snapshot versioning","Time-based key-value","Stream processing","Iterator design"],
}

ws = wb.create_sheet("Pattern Mastery", 1)
ws.sheet_properties.tabColor = C_PURPLE
N_PAT = len(PATTERNS)
MAIN_COLS = 10

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 1: TITLE
# ═══════════════════════════════════════════════════════════════════════════
ws.merge_cells("A1:J1")
cl = ws.cell(1,1); cl.value="PATTERN MASTERY — Interview Readiness Assessment"; cl.font=FT; cl.fill=FILL_BG; cl.alignment=AC
ws.merge_cells("A2:J2")
cl = ws.cell(2,1); cl.value='"If an interviewer gives me a brand-new problem from this pattern, can I derive the optimal solution without memorization?"'; cl.font=FM; cl.fill=FILL_BG; cl.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cl.font=Font("Calibri",italic=True,size=10,color=C_MUTED)
ws.row_dimensions[2].height = 28

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 2: STATISTICS DASHBOARD (row 4-6)
# ═══════════════════════════════════════════════════════════════════════════
r = 4
sec_title(ws, r, MAIN_COLS, "STATISTICS DASHBOARD")
r = 5
stat_labels = ["Total Patterns","Interview Ready","Needs Revision","Avg Recognition","Avg Confidence"]
for i, label in enumerate(stat_labels):
    c = 1 + i*2
    cl = ws.cell(r, c); cl.value=label; cl.font=F_STAT_L; cl.fill=FILL_C; cl.border=BD; cl.alignment=AC
    cl2 = ws.cell(r, c+1); cl2.fill=FILL_C; cl2.border=BD; cl2.alignment=AC

DATA_START = 10
DATA_END = DATA_START + N_PAT - 1
rec_range = f"C{DATA_START}:C{DATA_END}"
conf_range = f"H{DATA_START}:H{DATA_END}"

ws.cell(5,2).value = N_PAT; ws.cell(5,2).font = F_STAT_V
ws.cell(5,4).value = f'=COUNTIFS(C{DATA_START}:C{DATA_END},">=4",H{DATA_START}:H{DATA_END},">=4")'; ws.cell(5,4).font = F_STAT_V
ws.cell(5,6).value = f'={N_PAT}-B5'; ws.cell(5,6).font = F_STAT_V2
ws.cell(5,8).value = f'=IF(COUNTA(C{DATA_START}:C{DATA_END})>0,ROUND(AVERAGE(C{DATA_START}:C{DATA_END}),1),"—")'; ws.cell(5,8).font = F_STAT_V
ws.cell(5,10).value = f'=IF(COUNTA(H{DATA_START}:H{DATA_END})>0,ROUND(AVERAGE(H{DATA_START}:H{DATA_END}),1),"—")'; ws.cell(5,10).font = F_STAT_V

r = 6
more_stats = ["Strongest Pattern","Weakest Pattern","Completion %"]
for i, label in enumerate(more_stats):
    c = 1 + i*3
    cl = ws.cell(r, c); cl.value=label; cl.font=F_STAT_L; cl.fill=FILL_D; cl.border=BD; cl.alignment=AC

ws.cell(6,2).value = f'=IF(COUNTA(H{DATA_START}:H{DATA_END})>0,INDEX(A{DATA_START}:A{DATA_END},MATCH(MAX(H{DATA_START}:H{DATA_END}),H{DATA_START}:H{DATA_END},0)),"—")'
ws.cell(6,2).font = FG; ws.cell(6,2).fill=FILL_D; ws.cell(6,2).border=BD; ws.cell(6,2).alignment=AC
ws.merge_cells(start_row=6,start_column=2,end_row=6,end_column=3)

ws.cell(6,5).value = f'=IF(COUNTA(H{DATA_START}:H{DATA_END})>0,INDEX(A{DATA_START}:A{DATA_END},MATCH(MIN(H{DATA_START}:H{DATA_END}),H{DATA_START}:H{DATA_END},0)),"—")'
ws.cell(6,5).font = FR; ws.cell(6,5).fill=FILL_D; ws.cell(6,5).border=BD; ws.cell(6,5).alignment=AC
ws.merge_cells(start_row=6,start_column=5,end_row=6,end_column=6)

ws.cell(6,8).value = f'=IF(COUNTA(H{DATA_START}:H{DATA_END})>0,ROUND(COUNTIF(H{DATA_START}:H{DATA_END},">=4")/{N_PAT}*100,0)&"%","0%")'
ws.cell(6,8).font = F_STAT_V; ws.cell(6,8).fill=FILL_D; ws.cell(6,8).border=BD; ws.cell(6,8).alignment=AC
ws.merge_cells(start_row=6,start_column=8,end_row=6,end_column=9)

for c in range(1,MAIN_COLS+1):
    for rr in [5,6]:
        if not ws.cell(rr,c).fill or ws.cell(rr,c).fill.fgColor.rgb == "00000000":
            ws.cell(rr,c).fill = FILL_C if rr==5 else FILL_D
            ws.cell(rr,c).border = BD

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 3: LEGENDS (row 7-8)
# ═══════════════════════════════════════════════════════════════════════════
r = 8
sec_title(ws, r, MAIN_COLS, "RECOGNITION SCALE: 1=Never recognize  2=Need hints  3=Recognize after thinking  4=Recognize quickly  5=Instantly       CONFIDENCE: 1=Cannot solve  2=Need help  3=Struggle  4=Independent  5=Interview ready")

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 4: MAIN TRACKER TABLE (row 9 onwards)
# ═══════════════════════════════════════════════════════════════════════════
r = 9
main_headers = ["Pattern","Problems Solved","Recognition (1-5)","Can Solve Without Hints",
                "Can Explain Intuition","Can Explain Complexity","Can Solve Variants",
                "Confidence (1-5)","Weak Areas","Notes"]
main_widths = [26,15,16,20,18,20,16,15,28,28]
widths(ws, main_widths)
for i, h in enumerate(main_headers, 1):
    ws.cell(r, i, value=h)
hdr(ws, r, MAIN_COLS)

for idx, pat in enumerate(PATTERNS):
    rw = DATA_START + idx
    ws.cell(rw, 1, value=pat)
    body(ws, rw, MAIN_COLS, alt=idx%2==0)
    ws.cell(rw, 1).alignment = AL

pe = DATA_END
dv(ws, "C", DATA_START, pe, ["1","2","3","4","5"])
dv(ws, "D", DATA_START, pe, ["Yes","Partially","No"])
dv(ws, "E", DATA_START, pe, ["Yes","Partially","No"])
dv(ws, "F", DATA_START, pe, ["Yes","Partially","No"])
dv(ws, "G", DATA_START, pe, ["Yes","Partially","No"])
dv(ws, "H", DATA_START, pe, ["1","2","3","4","5"])

green_fill = PatternFill("solid", fgColor="1A3A2A")
yellow_fill = PatternFill("solid", fgColor="3A3520")
red_fill = PatternFill("solid", fgColor="3A1A1A")
green_font = Font("Calibri", bold=True, size=11, color=C_GREEN)
yellow_font = Font("Calibri", bold=True, size=11, color=C_YELLOW)
red_font = Font("Calibri", bold=True, size=11, color=C_RED)

for col in ["C", "H"]:
    rng = f"{col}{DATA_START}:{col}{pe}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"5"'], fill=green_fill, font=green_font))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"4"'], fill=green_fill, font=green_font))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"3"'], fill=yellow_fill, font=yellow_font))
    ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=['"1"','"2"'], fill=red_fill, font=red_font))

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 5: INTERVIEW DECISION (after main table)
# ═══════════════════════════════════════════════════════════════════════════
r = DATA_END + 2
sec_title(ws, r, MAIN_COLS, "INTERVIEW READINESS DECISION")
r += 1
dec_headers = ["Pattern","Recognition","Confidence","No Hints?","Explain?","Variants?","Status"]
for i, h in enumerate(dec_headers, 1):
    ws.cell(r, i, value=h)
hdr(ws, r, len(dec_headers))

for idx, pat in enumerate(PATTERNS):
    rw = r + 1 + idx
    pr = DATA_START + idx
    ws.cell(rw, 1, value=pat)
    ws.cell(rw, 2).value = f'=C{pr}'; ws.cell(rw, 3).value = f'=H{pr}'
    ws.cell(rw, 4).value = f'=D{pr}'; ws.cell(rw, 5).value = f'=E{pr}'
    ws.cell(rw, 6).value = f'=G{pr}'
    ws.cell(rw, 7).value = f'=IF(AND(C{pr}>=4,H{pr}>=4,D{pr}="Yes",E{pr}="Yes",G{pr}="Yes"),"Interview Ready","Needs Revision")'
    body(ws, rw, len(dec_headers), alt=idx%2==0)
    ws.cell(rw, 1).alignment = AL

dec_end = r + N_PAT
status_rng = f"G{r+1}:G{dec_end}"
ws.conditional_formatting.add(status_rng, CellIsRule(operator="equal", formula=['"Interview Ready"'], fill=green_fill, font=green_font))
ws.conditional_formatting.add(status_rng, CellIsRule(operator="equal", formula=['"Needs Revision"'], fill=yellow_fill, font=yellow_font))

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 6: PATTERN READINESS LEVELS (after decision)
# ═══════════════════════════════════════════════════════════════════════════
r = dec_end + 2
sec_title(ws, r, MAIN_COLS, "PATTERN READINESS LEVELS")
r += 1
rd_headers = ["Pattern","Beginner","Intermediate","Advanced","Interview Ready"]
for i, h in enumerate(rd_headers, 1):
    ws.cell(r, i, value=h)
hdr(ws, r, len(rd_headers))

for idx, pat in enumerate(PATTERNS):
    rw = r + 1 + idx
    ws.cell(rw, 1, value=pat)
    body(ws, rw, len(rd_headers), alt=idx%2==0)
    ws.cell(rw, 1).alignment = AL

rd_end = r + N_PAT
for col in ["B","C","D","E"]:
    dv(ws, col, r+1, rd_end, ["Done","In Progress","Not Started"])
    rng = f"{col}{r+1}:{col}{rd_end}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Done"'], fill=green_fill, font=green_font))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"In Progress"'], fill=yellow_fill, font=yellow_font))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Not Started"'], fill=red_fill, font=red_font))

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 7: PATTERN MASTERY CHECKLISTS
# ═══════════════════════════════════════════════════════════════════════════
r = rd_end + 2
sec_title(ws, r, MAIN_COLS, "PATTERN MASTERY CHECKLISTS — Sub-skill breakdown per pattern")
r += 1

for pat in PATTERNS:
    items = CHECKLISTS.get(pat, [])
    if not items:
        continue
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    cl = ws.cell(r, 1); cl.value=pat; cl.font=FS2; cl.fill=FILL_SEC; cl.alignment=AL; cl.border=BD
    for c in range(2,5):
        ws.cell(r,c).fill=FILL_SEC; ws.cell(r,c).border=BD
    r += 1

    chk_headers = ["#","Sub-skill","Mastered?","Notes"]
    for i,h in enumerate(chk_headers,1):
        ws.cell(r,i,value=h)
    hdr(ws,r,4)
    r += 1

    for j, item in enumerate(items):
        ws.cell(r, 1, value=j+1)
        ws.cell(r, 2, value=item)
        body(ws, r, 4, alt=j%2==0)
        ws.cell(r, 2).alignment = AL
        r += 1

    chk_start = r - len(items)
    chk_end = r - 1
    dv(ws, "C", chk_start, chk_end, ["Yes","Partially","No"])
    rng = f"C{chk_start}:C{chk_end}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Yes"'], fill=green_fill, font=green_font))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Partially"'], fill=yellow_fill, font=yellow_font))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"No"'], fill=red_fill, font=red_font))

    r += 1

ws.freeze_panes = "A10"
ws.sheet_view.showGridLines = False

wb.save(fp)
print(f"Added 'Pattern Mastery' sheet to: {fp}")
print(f"All sheets: {wb.sheetnames}")
print(f"Patterns: {N_PAT} | Checklist sub-skills: {sum(len(v) for v in CHECKLISTS.values())}")
