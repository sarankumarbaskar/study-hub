import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
from collections import Counter

wb = openpyxl.Workbook()

C_BG = "0A0E17"; C_DARK = "111827"; C_CARD = "1A2234"; C_BORDER = "2D3748"
C_TEXT = "E2E8F0"; C_MUTED = "94A3B8"; C_ACCENT = "60A5FA"; C_GREEN = "34D399"
C_YELLOW = "FBBF24"; C_ORANGE = "FB923C"; C_RED = "F87171"; C_PINK = "F472B6"
C_PURPLE = "A78BFA"; C_WHITE = "FFFFFF"; C_CYAN = "22D3EE"

FILL_H = PatternFill("solid", fgColor=C_ACCENT)
FILL_D = PatternFill("solid", fgColor=C_DARK)
FILL_C = PatternFill("solid", fgColor=C_CARD)
FILL_BG = PatternFill("solid", fgColor=C_BG)
FILL_SEC = PatternFill("solid", fgColor="1E3A5F")

FH = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
FT = Font(name="Calibri", bold=True, size=16, color=C_ACCENT)
FS = Font(name="Calibri", bold=True, size=12, color=C_TEXT)
FS2 = Font(name="Calibri", bold=True, size=11, color=C_ACCENT)
FB = Font(name="Calibri", size=11, color=C_TEXT)
FM = Font(name="Calibri", size=10, color=C_MUTED)
FL = Font(name="Calibri", size=11, color=C_ACCENT, underline="single")
FG = Font(name="Calibri", bold=True, size=11, color=C_GREEN)
FY = Font(name="Calibri", bold=True, size=11, color=C_YELLOW)
FR = Font(name="Calibri", bold=True, size=11, color=C_RED)
FP = Font(name="Calibri", bold=True, size=11, color=C_PURPLE)
FO = Font(name="Calibri", bold=True, size=11, color=C_ORANGE)
FC = Font(name="Calibri", bold=True, size=11, color=C_CYAN)

BD = Border(left=Side("thin", color=C_BORDER), right=Side("thin", color=C_BORDER),
            top=Side("thin", color=C_BORDER), bottom=Side("thin", color=C_BORDER))
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left", vertical="center", wrap_text=True)

def hdr(ws, r, n):
    for c in range(1, n+1):
        cl = ws.cell(row=r, column=c); cl.font=FH; cl.fill=FILL_H; cl.alignment=AC; cl.border=BD

def row_style(ws, r, n, alt=False):
    f = FILL_C if alt else FILL_D
    for c in range(1, n+1):
        cl = ws.cell(row=r, column=c); cl.font=FB; cl.fill=f; cl.alignment=AC if c!=2 else AL; cl.border=BD

def sec_row(ws, r, n, text):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n)
    cl = ws.cell(row=r, column=1); cl.value=text; cl.font=FS2; cl.fill=FILL_SEC; cl.alignment=AL; cl.border=BD

def dv(ws, col, r1, r2, opts):
    v = DataValidation(type="list", formula1=f'"{",".join(opts)}"', allow_blank=True)
    ws.add_data_validation(v); v.add(f"{col}{r1}:{col}{r2}")

def widths(ws, w_list):
    for i, w in enumerate(w_list, 1): ws.column_dimensions[get_column_letter(i)].width = w

PROBLEMS = [
    (1,"Two Sum",1,"Easy","Arrays & Hashing","Very High","Phase 1"),
    (2,"Group Anagrams",49,"Medium","Arrays & Hashing","Very High","Phase 1"),
    (3,"Product of Array Except Self",238,"Medium","Arrays & Hashing","Very High","Phase 1"),
    (4,"Longest Consecutive Sequence",128,"Medium","Arrays & Hashing","High","Phase 1"),
    (5,"Top K Frequent Elements",347,"Medium","Arrays & Hashing","Very High","Phase 1"),
    (135,"Rotate Image",48,"Medium","Arrays & Hashing","Very High","Phase 1"),
    (136,"Spiral Matrix",54,"Medium","Arrays & Hashing","Very High","Phase 1"),
    (6,"Container With Most Water",11,"Medium","Two Pointers","Very High","Phase 1"),
    (7,"3Sum",15,"Medium","Two Pointers","Very High","Phase 1"),
    (8,"Sort Colors",75,"Medium","Two Pointers","High","Phase 1"),
    (9,"Find the Duplicate Number",287,"Medium","Two Pointers","High","Phase 1"),
    (10,"Longest Palindromic Substring",5,"Medium","Two Pointers","Very High","Phase 1"),
    (11,"Longest Substring Without Repeating Chars",3,"Medium","Sliding Window","Very High","Phase 1"),
    (12,"Longest Repeating Character Replacement",424,"Medium","Sliding Window","High","Phase 1"),
    (13,"Permutation in String",567,"Medium","Sliding Window","High","Phase 1"),
    (14,"Find All Anagrams in a String",438,"Medium","Sliding Window","High","Phase 1"),
    (15,"Minimum Window Substring",76,"Hard","Sliding Window","Very High","Phase 1"),
    (16,"Fruit Into Baskets",904,"Medium","Sliding Window","Medium","Phase 1"),
    (17,"Max Consecutive Ones III",1004,"Medium","Sliding Window","Medium","Phase 1"),
    (18,"Subarray Product Less Than K",713,"Medium","Sliding Window","Medium","Phase 1"),
    (19,"Frequency of Most Frequent Element",1838,"Medium","Sliding Window","Medium","Phase 1"),
    (20,"Sliding Window Maximum",239,"Hard","Sliding Window","Very High","Phase 1"),
    (137,"Best Time to Buy and Sell Stock",121,"Easy","Sliding Window","Very High","Phase 1"),
    (21,"Search in Rotated Sorted Array",33,"Medium","Binary Search","Very High","Phase 1"),
    (22,"Find Min in Rotated Sorted Array",153,"Medium","Binary Search","High","Phase 1"),
    (23,"Find Peak Element",162,"Medium","Binary Search","High","Phase 1"),
    (24,"Koko Eating Bananas",875,"Medium","Binary Search","Very High","Phase 1"),
    (25,"Capacity To Ship Packages",1011,"Medium","Binary Search","High","Phase 1"),
    (26,"Min Days to Make m Bouquets",1482,"Medium","Binary Search","Medium","Phase 1"),
    (27,"Split Array Largest Sum",410,"Hard","Binary Search","High","Phase 1"),
    (28,"Find First and Last Position",34,"Medium","Binary Search","Very High","Phase 1"),
    (29,"Median of Two Sorted Arrays",4,"Hard","Binary Search","Very High","Phase 1"),
    (30,"Kth Smallest in Sorted Matrix",378,"Medium","Binary Search","High","Phase 1"),
    (31,"Single Number",136,"Easy","Bit Manipulation","Very High","Phase 1"),
    (32,"Number of 1 Bits",191,"Easy","Bit Manipulation","High","Phase 1"),
    (33,"Counting Bits",338,"Easy","Bit Manipulation","Medium","Phase 1"),
    (34,"Sum of Two Integers",371,"Medium","Bit Manipulation","Medium","Phase 1"),
    (35,"Daily Temperatures",739,"Medium","Stack","Very High","Phase 2"),
    (36,"Next Greater Element II",503,"Medium","Stack","Medium","Phase 2"),
    (37,"Decode String",394,"Medium","Stack","High","Phase 2"),
    (38,"Asteroid Collision",735,"Medium","Stack","High","Phase 2"),
    (39,"Remove K Digits",402,"Medium","Stack","High","Phase 2"),
    (40,"Largest Rectangle in Histogram",84,"Hard","Stack","Very High","Phase 2"),
    (41,"Trapping Rain Water",42,"Hard","Stack","Very High","Phase 2"),
    (42,"Sum of Subarray Minimums",907,"Medium","Stack","Medium","Phase 2"),
    (43,"Kth Largest Element in Array",215,"Medium","Heap","Very High","Phase 2"),
    (44,"Top K Frequent Elements (Heap)",347,"Medium","Heap","Very High","Phase 2"),
    (45,"K Closest Points to Origin",973,"Medium","Heap","High","Phase 2"),
    (46,"Find Median from Data Stream",295,"Hard","Heap","Very High","Phase 2"),
    (47,"Meeting Rooms II",253,"Medium","Heap","Very High","Phase 2"),
    (48,"Merge K Sorted Lists",23,"Hard","Heap","Very High","Phase 2"),
    (49,"Smallest Range Covering K Lists",632,"Hard","Heap","Medium","Phase 2"),
    (50,"Reverse Linked List",206,"Easy","Linked List","Very High","Phase 2"),
    (51,"Linked List Cycle",141,"Easy","Linked List","Very High","Phase 2"),
    (52,"Reorder List",143,"Medium","Linked List","High","Phase 2"),
    (53,"Remove Nth Node From End",19,"Medium","Linked List","Very High","Phase 2"),
    (54,"Copy List with Random Pointer",138,"Medium","Linked List","High","Phase 2"),
    (55,"Add Two Numbers",2,"Medium","Linked List","Very High","Phase 2"),
    (56,"Reverse Nodes in k-Group",25,"Hard","Linked List","High","Phase 2"),
    (57,"Merge Intervals",56,"Medium","Intervals","Very High","Phase 2"),
    (58,"Insert Interval",57,"Medium","Intervals","High","Phase 2"),
    (59,"Non-overlapping Intervals",435,"Medium","Intervals","High","Phase 2"),
    (60,"Meeting Rooms",252,"Easy","Intervals","High","Phase 2"),
    (61,"BT Level Order Traversal",102,"Medium","Trees","Very High","Phase 3"),
    (62,"BT Right Side View",199,"Medium","Trees","Very High","Phase 3"),
    (63,"Validate BST",98,"Medium","Trees","Very High","Phase 3"),
    (64,"Kth Smallest in BST",230,"Medium","Trees","High","Phase 3"),
    (65,"Construct BT from Preorder+Inorder",105,"Medium","Trees","High","Phase 3"),
    (66,"Diameter of Binary Tree",543,"Easy","Trees","Very High","Phase 3"),
    (67,"Balanced Binary Tree",110,"Easy","Trees","High","Phase 3"),
    (68,"LCA of a BST",235,"Medium","Trees","Very High","Phase 3"),
    (69,"LCA of a Binary Tree",236,"Medium","Trees","Very High","Phase 3"),
    (70,"House Robber III",337,"Medium","Trees","Medium","Phase 3"),
    (71,"BT Maximum Path Sum",124,"Hard","Trees","Very High","Phase 3"),
    (72,"Serialize/Deserialize BT",297,"Hard","Trees","Very High","Phase 3"),
    (73,"All Nodes Distance K in BT",863,"Medium","Trees","High","Phase 3"),
    (74,"Find Duplicate Subtrees",652,"Medium","Trees","Medium","Phase 3"),
    (75,"Flatten BT to Linked List",114,"Medium","Trees","High","Phase 3"),
    (76,"Number of Islands",200,"Medium","Graphs","Very High","Phase 3"),
    (77,"Max Area of Island",695,"Medium","Graphs","High","Phase 3"),
    (78,"Rotting Oranges",994,"Medium","Graphs","Very High","Phase 3"),
    (79,"Pacific Atlantic Water Flow",417,"Medium","Graphs","High","Phase 3"),
    (80,"Course Schedule",207,"Medium","Graphs","Very High","Phase 3"),
    (81,"Course Schedule II",210,"Medium","Graphs","Very High","Phase 3"),
    (82,"Clone Graph",133,"Medium","Graphs","High","Phase 3"),
    (83,"Number of Provinces",547,"Medium","Graphs","High","Phase 3"),
    (84,"Redundant Connection",684,"Medium","Graphs","Medium","Phase 3"),
    (85,"Accounts Merge",721,"Medium","Graphs","High","Phase 3"),
    (86,"Network Delay Time",743,"Medium","Graphs","High","Phase 3"),
    (87,"Cheapest Flights K Stops",787,"Medium","Graphs","High","Phase 3"),
    (88,"Path With Minimum Effort",1631,"Medium","Graphs","Medium","Phase 3"),
    (89,"Word Ladder",127,"Hard","Graphs","Very High","Phase 3"),
    (90,"Graph Valid Tree",261,"Medium","Graphs","High","Phase 3"),
    (91,"Critical Connections",1192,"Hard","Graphs","Medium","Phase 3"),
    (92,"Min Cost Connect All Points",1584,"Medium","Graphs","Medium","Phase 3"),
    (93,"Connected Components",323,"Medium","Graphs","High","Phase 3"),
    (94,"Find Eventual Safe States",802,"Medium","Graphs","Medium","Phase 3"),
    (95,"Alien Dictionary",269,"Hard","Graphs","Very High","Phase 3"),
    (96,"House Robber",198,"Medium","DP","Very High","Phase 4"),
    (97,"House Robber II",213,"Medium","DP","High","Phase 4"),
    (98,"Decode Ways",91,"Medium","DP","Very High","Phase 4"),
    (99,"Coin Change",322,"Medium","DP","Very High","Phase 4"),
    (100,"Coin Change II",518,"Medium","DP","High","Phase 4"),
    (101,"Partition Equal Subset Sum",416,"Medium","DP","High","Phase 4"),
    (102,"Target Sum",494,"Medium","DP","High","Phase 4"),
    (103,"Word Break",139,"Medium","DP","Very High","Phase 4"),
    (104,"Longest Common Subsequence",1143,"Medium","DP","Very High","Phase 4"),
    (105,"Edit Distance",72,"Medium","DP","Very High","Phase 4"),
    (106,"Longest Palindromic Subsequence",516,"Medium","DP","Medium","Phase 4"),
    (107,"Unique Paths",62,"Medium","DP","Very High","Phase 4"),
    (108,"Minimum Path Sum",64,"Medium","DP","High","Phase 4"),
    (109,"Maximal Square",221,"Medium","DP","High","Phase 4"),
    (110,"Longest Increasing Subsequence",300,"Medium","DP","Very High","Phase 4"),
    (111,"Russian Doll Envelopes",354,"Hard","DP","Medium","Phase 4"),
    (112,"Buy Stock III",123,"Hard","DP","High","Phase 4"),
    (113,"Buy Stock Cooldown",309,"Medium","DP","High","Phase 4"),
    (114,"Burst Balloons",312,"Hard","DP","High","Phase 4"),
    (115,"Different Ways Add Parentheses",241,"Medium","DP","Medium","Phase 4"),
    (138,"Maximum Product Subarray",152,"Medium","DP","Very High","Phase 4"),
    (139,"Longest Increasing Path Matrix",329,"Hard","DP","Very High","Phase 4"),
    (116,"Jump Game",55,"Medium","Greedy","Very High","Phase 4"),
    (117,"Jump Game II",45,"Medium","Greedy","Very High","Phase 4"),
    (118,"Gas Station",134,"Medium","Greedy","High","Phase 4"),
    (119,"Task Scheduler",621,"Medium","Greedy","Very High","Phase 4"),
    (140,"Maximum Subarray",53,"Medium","Greedy","Very High","Phase 4"),
    (120,"Combination Sum",39,"Medium","Backtracking","Very High","Phase 4"),
    (121,"Permutations",46,"Medium","Backtracking","Very High","Phase 4"),
    (122,"Generate Parentheses",22,"Medium","Backtracking","Very High","Phase 4"),
    (123,"Word Search",79,"Medium","Backtracking","Very High","Phase 4"),
    (124,"N-Queens",51,"Hard","Backtracking","High","Phase 4"),
    (141,"Subsets",78,"Medium","Backtracking","Very High","Phase 4"),
    (125,"LRU Cache",146,"Medium","Design","Very High","Phase 5"),
    (126,"LFU Cache",460,"Hard","Design","High","Phase 5"),
    (127,"Implement Trie",208,"Medium","Design","Very High","Phase 5"),
    (128,"Design Add and Search Words",211,"Medium","Design","High","Phase 5"),
    (129,"Time Based Key-Value Store",981,"Medium","Design","High","Phase 5"),
    (130,"Design HashMap",706,"Easy","Design","High","Phase 5"),
    (131,"Insert Delete GetRandom O(1)",380,"Medium","Design","Very High","Phase 5"),
    (132,"Snapshot Array",1146,"Medium","Design","Medium","Phase 5"),
    (133,"Stock Price Fluctuation",2034,"Medium","Design","Medium","Phase 5"),
    (134,"Design Search Autocomplete",642,"Hard","Design","High","Phase 5"),
    (142,"Word Search II",212,"Hard","Design","Very High","Phase 5"),
]

today = datetime.now().date()

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.active; ws.title = "Dashboard"; ws.sheet_properties.tabColor = C_ACCENT
widths(ws, [26,14,14,14,14,14,14,14,14,14,14])

ws.merge_cells("A1:K1"); c=ws["A1"]; c.value="INTERVIEW PREPARATION TRACKER"; c.font=FT; c.fill=FILL_BG; c.alignment=AC
ws.merge_cells("A2:K2"); c=ws["A2"]; c.value="5 YoE Java Backend @ Red Hat  |  Target: SDE-2 / Senior SWE  |  40+ LPA  |  8-Month Plan"; c.font=FM; c.fill=FILL_BG; c.alignment=AC

r=4; ws.cell(r,1,value="PHASE 1: DSA-HEAVY (Months 1–4) — DSA is the gating round at Tier-1 companies").font=FS; ws.cell(r,1).fill=FILL_BG
p1=[("DSA","45%","~6.75h","142 problems — highest interview impact"),("System Design + DDIA","25%","~3.75h","DDIA Ch 1-6 + 3 HLD designs"),("Java & Backend","15%","~2.25h","Effective Java + Design Patterns started"),("Kafka & Redis","10%","~1.5h","Kafka fundamentals + consumers"),("Cloud & Platform","5%","~0.75h","Docker basics")]
for i,h in enumerate(["Track","Weight","Hrs/Wk","Focus"]): ws.cell(5,1+i,value=h)
hdr(ws,5,4)
for j,(t,w,h,s) in enumerate(p1):
    rw=6+j
    for ci,v in enumerate([t,w,h,s],1): ws.cell(rw,ci,value=v)
    row_style(ws,rw,4,j%2==0)

r=12; ws.cell(r,1,value="PHASE 2: BALANCED (Months 5–8) — After DSA patterns are solid").font=FS; ws.cell(r,1).fill=FILL_BG
p2=[("DSA","20%","~3h","Revision + hard problems + mock interviews"),("System Design + DDIA","30%","~4.5h","DDIA Ch 7-9 + 5 HLD designs + mocks"),("Java & Backend","25%","~3.75h","Design Patterns done + Spring Boot deep dive"),("Kafka & Redis","15%","~2.25h","Kafka reliability + Redis patterns"),("Cloud & Platform","10%","~1.5h","K8s, OpenShift, CI/CD, AWS")]
for i,h in enumerate(["Track","Weight","Hrs/Wk","Focus"]): ws.cell(13,1+i,value=h)
hdr(ws,13,4)
for j,(t,w,h,s) in enumerate(p2):
    rw=14+j
    for ci,v in enumerate([t,w,h,s],1): ws.cell(rw,ci,value=v)
    row_style(ws,rw,4,j%2==0)

r=20; ws.cell(r,1,value="MONTHLY MILESTONES").font=FS; ws.cell(r,1).fill=FILL_BG
mm_h=["Month","DSA","Java & Backend","System Design","Kafka/Redis","Cloud/Platform"]
for i,h in enumerate(mm_h): ws.cell(21,1+i,value=h)
hdr(ws,21,6)
milestones=[
    ("Month 1","Phase 1: Arrays, Two Pointers, SW","Effective Java started","DDIA Ch 1-2","Kafka fundamentals","Docker basics"),
    ("Month 2","Phase 1-2: BS, Stack, Heap","Effective Java 50%","DDIA Ch 3-4","Kafka consumers","Docker multi-stage"),
    ("Month 3","Phase 2: LL, Intervals","Effective Java done","DDIA Ch 5 + URL Shortener","Kafka reliability","K8s basics"),
    ("Month 4","Phase 3: Trees","Design Patterns started","DDIA Ch 6 + Rate Limiter","Redis fundamentals","K8s probes/HPA"),
    ("Month 5","Phase 3: Graphs","Design Patterns 50%","DDIA Ch 7 + Notification Sys","Redis locking","OpenShift basics"),
    ("Month 6","Phase 4: DP","Design Patterns done","DDIA Ch 8-9 + Inventory","Integration practice","CI/CD pipelines"),
    ("Month 7","Phase 4-5: Greedy, BT, Design","Spring Boot deep dive","Payment Sys + Search","E2E design","AWS fundamentals"),
    ("Month 8","Review + Mocks","Spring Security","Dist Cache + Review","Mock interviews","Mock interviews"),
]
for j,row_data in enumerate(milestones):
    rw=22+j
    for ci,v in enumerate(row_data,1): ws.cell(rw,ci,value=v)
    row_style(ws,rw,6,j%2==0); ws.cell(rw,1).font=FS2

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2: DSA PROBLEMS
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("DSA Problems"); ws2.sheet_properties.tabColor = C_PINK
h2=["#","Problem","LC#","Diff","Pattern","Phase","Freq","Status","Date Attempted","Date Solved","Time(min)","Approach","1st Try?","Confidence","Rev1","Rev2","Rev3","URL","Notes"]
w2=[5,36,6,9,17,9,11,12,13,13,10,13,9,11,11,11,11,40,28]
widths(ws2,w2)
for i,h in enumerate(h2,1): ws2.cell(1,i,value=h)
hdr(ws2,1,len(h2)); ws2.freeze_panes="A2"
ws2.auto_filter.ref=f"A1:{get_column_letter(len(h2))}{len(PROBLEMS)+1}"
for idx,(num,name,lc,diff,pat,freq,ph) in enumerate(PROBLEMS):
    r=idx+2
    url=f"https://leetcode.com/problems/{name.lower().replace(' ','-').replace('(','').replace(')','').replace(',','')}/"
    for ci,v in enumerate([num,name,lc,diff,pat,ph,freq,"","","","","","","","","","",url,""],1): ws2.cell(r,ci,value=v)
    row_style(ws2,r,len(h2),idx%2==0)
    ws2.cell(r,18).font=FL
    d=ws2.cell(r,4)
    if diff=="Easy": d.font=FG
    elif diff=="Medium": d.font=FY
    else: d.font=FR
de=len(PROBLEMS)+1
dv(ws2,"H",2,de,["Not Started","Attempted","Solved","Revisit","Mastered"])
dv(ws2,"L",2,de,["Brute Force","Optimal","Saw Solution","Hint Used"])
dv(ws2,"M",2,de,["Yes","No"])
dv(ws2,"N",2,de,["1","2","3","4","5"])

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3: JAVA & BACKEND
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Java & Backend"); ws3.sheet_properties.tabColor = C_ORANGE
h3=["#","Topic","Sub-Track","Category","Status","Date Started","Date Done","Confidence","Can Explain in Interview?","Notes"]
w3=[4,38,18,16,13,13,13,12,18,30]
widths(ws3,w3)
for i,h in enumerate(h3,1): ws3.cell(1,i,value=h)
hdr(ws3,1,len(h3)); ws3.freeze_panes="A2"

java_topics = [
    ("Effective Java","Generics","Bounded wildcards, type erasure, PECS"),
    ("Effective Java","Collections","Choosing the right collection, immutable collections"),
    ("Effective Java","Enums","Enum with behavior, EnumSet, EnumMap"),
    ("Effective Java","Lambdas & Streams","Functional interfaces, stream pipeline, collectors"),
    ("Effective Java","Immutability","Immutable classes, defensive copies"),
    ("Effective Java","equals/hashCode","Contract, consistent implementation"),
    ("Effective Java","Best Practices","Builder, try-with-resources, Optional, method design"),
    ("Design Patterns","Strategy","Swap algorithms at runtime"),
    ("Design Patterns","Observer","Event-driven notifications"),
    ("Design Patterns","Factory","Decouple object creation"),
    ("Design Patterns","Singleton","Exactly one instance"),
    ("Design Patterns","Command","Encapsulate requests as objects"),
    ("Design Patterns","Adapter","Incompatible interface bridging"),
    ("Design Patterns","Decorator","Add behavior dynamically"),
    ("Spring Boot Core","Dependency Injection","Constructor vs field vs setter, @Qualifier, @Primary"),
    ("Spring Boot Core","Bean Lifecycle","@PostConstruct, @PreDestroy, scopes"),
    ("Spring Boot Core","Auto Configuration","@SpringBootApplication, conditional beans"),
    ("Spring Boot Core","Profiles","Environment-specific config, application-{profile}.yml"),
    ("Spring Boot Data","Transactions","@Transactional, propagation, isolation, rollback"),
    ("Spring Boot Data","JPA Basics","Entity mapping, relationships, N+1 problem"),
    ("Spring Security","Authentication","Form login, Basic auth, UserDetailsService"),
    ("Spring Security","Authorization","Role-based, @PreAuthorize, method-level"),
    ("Spring Security","OAuth2","Authorization code flow, resource server"),
    ("Spring Security","JWT","Token structure, signing, stateless auth, refresh"),
]
for idx,(sub,topic,cat) in enumerate(java_topics):
    r=idx+2
    for ci,v in enumerate([idx+1,topic,sub,cat,"","","","","",""],1): ws3.cell(r,ci,value=v)
    row_style(ws3,r,len(h3),idx%2==0)
je=len(java_topics)+1
dv(ws3,"E",2,je,["Not Started","Reading","Practicing","Done","Revision Needed"])
dv(ws3,"H",2,je,["1","2","3","4","5"])
dv(ws3,"I",2,je,["Yes","Partially","No"])

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4: SYSTEM DESIGN
# ═══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("System Design"); ws4.sheet_properties.tabColor = C_ACCENT
h4=["#","System / Chapter","Track","Key Concepts","Status","Date Started","Date Done","Confidence","Doc Written?","Mock Practiced?","Notes"]
w4=[4,30,14,45,13,13,13,12,12,13,28]
widths(ws4,w4)
for i,h in enumerate(h4,1): ws4.cell(1,i,value=h)
hdr(ws4,1,len(h4)); ws4.freeze_panes="A2"

sd_topics = [
    ("DDIA Ch1: Reliability, Scalability","DDIA","Reliability vs availability, latency percentiles, load parameters"),
    ("DDIA Ch2: Data Models & Query Langs","DDIA","Relational vs document vs graph, schema-on-read vs write"),
    ("DDIA Ch3: Storage and Retrieval","DDIA","LSM-trees vs B-trees, SSTables, write amplification"),
    ("DDIA Ch4: Encoding and Evolution","DDIA","Avro/Protobuf, schema evolution, backward/forward compat"),
    ("DDIA Ch5: Replication","DDIA","Leader-follower, multi-leader, leaderless, quorum"),
    ("DDIA Ch6: Partitioning","DDIA","Hash vs range, rebalancing, secondary indexes"),
    ("DDIA Ch7: Transactions","DDIA","ACID, isolation levels, serializability, 2PC"),
    ("DDIA Ch8: Trouble with Dist Systems","DDIA","Partial failures, unreliable clocks, Byzantine faults"),
    ("DDIA Ch9: Consistency and Consensus","DDIA","Linearizability, causality, Raft/Paxos"),
    ("Design: URL Shortener","HLD","Hashing, base62, read-heavy, caching"),
    ("Design: Rate Limiter","HLD","Token bucket, sliding window, distributed"),
    ("Design: Notification System","HLD","Push/pull, fanout, message queues, delivery guarantees"),
    ("Design: Inventory Service","HLD","Consistency, distributed locking, stock reservation"),
    ("Design: Order Management","HLD","Saga pattern, state machines, idempotency"),
    ("Design: Payment System","HLD","Exactly-once, reconciliation, ledger design"),
    ("Design: Search Service","HLD","Inverted index, Elasticsearch, ranking, autocomplete"),
    ("Design: Distributed Cache","HLD","Consistent hashing, eviction, cache coherence"),
]
for idx,(topic,track,concepts) in enumerate(sd_topics):
    r=idx+2
    for ci,v in enumerate([idx+1,topic,track,concepts,"","","","","","",""],1): ws4.cell(r,ci,value=v)
    row_style(ws4,r,len(h4),idx%2==0)
se=len(sd_topics)+1
dv(ws4,"E",2,se,["Not Started","Reading","Practicing","Done","Revision Needed"])
dv(ws4,"H",2,se,["1","2","3","4","5"])
dv(ws4,"I",2,se,["Yes","No"])
dv(ws4,"J",2,se,["Yes","No"])

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5: KAFKA & REDIS
# ═══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Kafka & Redis"); ws5.sheet_properties.tabColor = C_RED
h5=["#","Topic","Technology","Category","Key Concepts","Status","Confidence","Can Explain?","Notes"]
w5=[4,28,10,16,48,13,12,12,28]
widths(ws5,w5)
for i,h in enumerate(h5,1): ws5.cell(1,i,value=h)
hdr(ws5,1,len(h5)); ws5.freeze_panes="A2"

kr_topics = [
    ("Topics","Kafka","Fundamentals","Partitioning strategy, naming conventions, compacted topics"),
    ("Brokers","Kafka","Fundamentals","Leader election, ISR, controller"),
    ("Partitions","Kafka","Fundamentals","Key-based routing, ordering, partition count tradeoffs"),
    ("Replication","Kafka","Fundamentals","Replication factor, acks=all vs acks=1, min.insync.replicas"),
    ("Consumer Groups","Kafka","Consumers","Parallel consumption, group coordinator, partition assignment"),
    ("Rebalancing","Kafka","Consumers","Eager vs cooperative, static membership"),
    ("Offset Management","Kafka","Consumers","Auto vs manual commit, at-least-once vs at-most-once"),
    ("Retry","Kafka","Reliability","Retry topics, exponential backoff, max retries"),
    ("DLQ","Kafka","Reliability","Dead letter queue design, monitoring, reprocessing"),
    ("Idempotency","Kafka","Reliability","Producer idempotency, PID/sequence"),
    ("Exactly Once","Kafka","Reliability","Transactional producer, read-process-write, EOS"),
    ("Cache Aside","Redis","Patterns","Read-through, lazy population, consistency tradeoffs"),
    ("TTL","Redis","Patterns","Time-based expiry, volatile vs non-volatile, memory mgmt"),
    ("Cache Stampede","Redis","Patterns","Thundering herd, probabilistic early expiration, locking"),
    ("Distributed Locking","Redis","Patterns","Redlock, SETNX + TTL, fencing tokens"),
    ("Rate Limiting","Redis","Patterns","Token bucket with Redis, sliding window with sorted sets"),
]
for idx,(topic,tech,cat,concepts) in enumerate(kr_topics):
    r=idx+2
    for ci,v in enumerate([idx+1,topic,tech,cat,concepts,"","","",""],1): ws5.cell(r,ci,value=v)
    row_style(ws5,r,len(h5),idx%2==0)
ke=len(kr_topics)+1
dv(ws5,"F",2,ke,["Not Started","Learning","Practicing","Done","Revision Needed"])
dv(ws5,"G",2,ke,["1","2","3","4","5"])
dv(ws5,"H",2,ke,["Yes","Partially","No"])

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6: CLOUD & PLATFORM
# ═══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Cloud & Platform"); ws6.sheet_properties.tabColor = C_GREEN
h6=["#","Topic","Technology","Key Concepts","Status","Confidence","Hands-On Done?","Notes"]
w6=[4,28,14,48,13,12,13,28]
widths(ws6,w6)
for i,h in enumerate(h6,1): ws6.cell(1,i,value=h)
hdr(ws6,1,len(h6)); ws6.freeze_panes="A2"

cloud_topics = [
    ("Images & Dockerfile","Docker","Base images, layer caching, image size optimization"),
    ("Containers & Lifecycle","Docker","Process isolation, resource limits, restart policies"),
    ("Volumes & Networks","Docker","Bind mounts, named volumes, bridge/overlay networks"),
    ("Multi-stage Builds","Docker","Separate build/runtime stages, minimal prod images"),
    ("Pods","Kubernetes","Pod lifecycle, init containers, sidecar pattern"),
    ("Deployments","Kubernetes","Rolling updates, rollback, replica management"),
    ("Services","Kubernetes","ClusterIP, NodePort, LoadBalancer, service discovery"),
    ("ConfigMaps & Secrets","Kubernetes","Externalized config, sensitive data, encryption"),
    ("HPA","Kubernetes","CPU/memory-based scaling, custom metrics"),
    ("Probes","Kubernetes","Readiness, liveness, startup probes"),
    ("Routes","OpenShift","Ingress routing, TLS termination, path-based routing"),
    ("Security Contexts","OpenShift","SCCs, pod security, non-root containers"),
    ("OpenShift Pipelines","OpenShift","Tekton CI/CD, PipelineRun, TaskRun"),
    ("Pipelines & Stages","CI/CD","gitlab-ci.yml, triggers, artifacts, parallel jobs"),
    ("Deployment Strategies","CI/CD","Blue-green, canary, rollback strategies"),
    ("EC2 & ECS","AWS","Instance types, ASG, Fargate vs EC2 launch"),
    ("Lambda","AWS","Event-driven, cold starts, concurrency limits"),
    ("S3","AWS","Storage classes, lifecycle policies, presigned URLs"),
    ("RDS & DynamoDB","AWS","Multi-AZ, read replicas, GSI/LSI, single-table design"),
    ("SQS & SNS","AWS","Standard vs FIFO, visibility timeout, fanout pattern"),
    ("VPC & Load Balancer","AWS","Subnets, security groups, ALB vs NLB, path routing"),
    ("CloudWatch","AWS","Metrics, alarms, log groups, dashboards"),
]
for idx,(topic,tech,concepts) in enumerate(cloud_topics):
    r=idx+2
    for ci,v in enumerate([idx+1,topic,tech,concepts,"","","",""],1): ws6.cell(r,ci,value=v)
    row_style(ws6,r,len(h6),idx%2==0)
ce=len(cloud_topics)+1
dv(ws6,"E",2,ce,["Not Started","Learning","Practicing","Done","Revision Needed"])
dv(ws6,"F",2,ce,["1","2","3","4","5"])
dv(ws6,"G",2,ce,["Yes","No"])

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 7: DAILY LOG
# ═══════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Daily Log"); ws7.sheet_properties.tabColor = C_CYAN
h7=["Date","Day","Track","Activity","Problem / Topic","Time(min)","Outcome","Key Learning","Mood(1-5)","Cumulative Hours"]
w7=[12,8,17,20,35,10,12,40,10,14]
widths(ws7,w7)
for i,h in enumerate(h7,1): ws7.cell(1,i,value=h)
hdr(ws7,1,len(h7)); ws7.freeze_panes="A2"
for d in range(180):
    dt=today+timedelta(days=d); r=d+2
    ws7.cell(r,1,value=dt.strftime("%Y-%m-%d"))
    ws7.cell(r,2,value=dt.strftime("%a"))
    row_style(ws7,r,len(h7),d%2==0)
    if dt.strftime("%a") in ("Sat","Sun"): ws7.cell(r,2).font=FG
dv(ws7,"C",2,181,["DSA","Java & Backend","System Design","DDIA","Kafka","Redis","Docker","K8s","AWS","CI/CD","Blog Reading","Mock Interview"])
dv(ws7,"D",2,181,["New Problem","Revision","Pattern Study","Book Reading","Video Course","SD Practice","Mock Interview","Blog","Coding Practice","Hands-On Lab"])
dv(ws7,"G",2,181,["Solved","Partial","Failed","Completed","In Progress"])
dv(ws7,"I",2,181,["1","2","3","4","5"])

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 8: WEEKLY REVIEW
# ═══════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Weekly Review"); ws8.sheet_properties.tabColor = C_PURPLE
h8=["Week","Start","End","DSA New","DSA Rev","DSA Hrs","Java Hrs","SD Hrs","Kafka/Redis Hrs","Cloud Hrs","Total Hrs","Blogs","Mock Interviews","Best Win","Biggest Struggle","Next Week Focus"]
w8=[6,11,11,9,9,9,10,9,13,10,10,8,13,28,28,28]
widths(ws8,w8)
for i,h in enumerate(h8,1): ws8.cell(1,i,value=h)
hdr(ws8,1,len(h8)); ws8.freeze_panes="A2"
for w in range(32):
    r=w+2; s=today+timedelta(weeks=w); e=s+timedelta(days=6)
    ws8.cell(r,1,value=w+1); ws8.cell(r,2,value=s.strftime("%m/%d")); ws8.cell(r,3,value=e.strftime("%m/%d"))
    row_style(ws8,r,len(h8),w%2==0); ws8.cell(r,1).font=FS2

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 9: MOCK INTERVIEWS
# ═══════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Mock Interviews"); ws9.sheet_properties.tabColor = C_ORANGE
h9=["Date","Type","Platform","Company","Q1","Pattern","Diff","Solved?","Time","Q2","Pattern","Diff","Solved?","Time","Score(1-10)","Mistakes","Lessons"]
w9=[11,13,13,14,28,13,9,9,8,28,13,9,9,8,12,32,32]
widths(ws9,w9)
for i,h in enumerate(h9,1): ws9.cell(1,i,value=h)
hdr(ws9,1,len(h9)); ws9.freeze_panes="A2"
for r in range(2,32): row_style(ws9,r,len(h9),r%2==0)
dv(ws9,"B",2,31,["Timed Solo","Peer Mock","Platform","Company Prep","System Design Mock"])
dv(ws9,"C",2,31,["LeetCode","Pramp","Interviewing.io","Peer","Self","Hello Interview"])

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 10: BOOKS & BLOGS
# ═══════════════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("Books & Blogs"); ws10.sheet_properties.tabColor = C_YELLOW
h10=["#","Book / Resource","Author","Track","Target","Total Units","Done Units","% Done","Current Position","Start Date","Last Date","Notes"]
w10=[4,32,20,17,14,12,11,9,22,12,12,28]
widths(ws10,w10)
for i,h in enumerate(h10,1): ws10.cell(1,i,value=h)
hdr(ws10,1,len(h10))

books=[
    ("Effective Java","Joshua Bloch","Java & Backend","Month 3","90 items"),
    ("Head First Design Patterns","Various","Java & Backend","Month 4","14 chapters"),
    ("DDIA (Ch 1-9)","Martin Kleppmann","Distributed Systems","Month 6","9 chapters"),
    ("System Design Interview","Alex Xu","System Design","Month 5","16 chapters"),
    ("Docker Deep Dive","Nigel Poulton","Cloud/Platform","Month 4","20 chapters"),
    ("Spring Start Here","Laurentiu Spilca","Java & Backend","Month 5","14 chapters"),
    ("Spring Security in Action","Laurentiu Spilca","Java & Backend","Month 6","20 chapters"),
]
for idx,(t,a,tr,tgt,u) in enumerate(books):
    r=idx+2
    for ci,v in enumerate([idx+1,t,a,tr,tgt,u,"","","","","",""],1): ws10.cell(r,ci,value=v)
    row_style(ws10,r,len(h10),idx%2==0)

r=len(books)+3
ws10.merge_cells(start_row=r,start_column=1,end_row=r,end_column=len(h10))
ws10.cell(r,1,value="ENGINEERING BLOGS — Track 2-3 articles/week").font=FS2; ws10.cell(r,1).fill=FILL_SEC

blogs_h=["#","Blog","Focus","URL","Articles Read This Month","Last Read Date","Favorite Article","Notes"]
bw=[4,20,30,38,20,14,30,28]
r+=1
for i,h in enumerate(blogs_h,1): ws10.cell(r,i,value=h); ws10.column_dimensions[get_column_letter(i)].width=max(ws10.column_dimensions[get_column_letter(i)].width,bw[i-1])
hdr(ws10,r,len(blogs_h))

eng_blogs=[
    ("Uber Engineering","Microservices, real-time systems, scale","https://eng.uber.com/"),
    ("Netflix Tech Blog","Resilience, chaos engineering, streaming","https://netflixtechblog.com/"),
    ("Cloudflare Blog","Networking, edge computing, security","https://blog.cloudflare.com/"),
    ("Stripe Engineering","Payment systems, API design, reliability","https://stripe.com/blog/engineering"),
    ("LinkedIn Engineering","Data infrastructure, distributed systems","https://engineering.linkedin.com/blog"),
    ("Confluent Blog","Kafka, event streaming, real-time data","https://www.confluent.io/blog/"),
]
for idx,(name,focus,url) in enumerate(eng_blogs):
    rw=r+1+idx
    for ci,v in enumerate([idx+1,name,focus,url,"","","",""],1): ws10.cell(rw,ci,value=v)
    row_style(ws10,rw,len(blogs_h),idx%2==0)
    ws10.cell(rw,4).font=FL

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 11: APPLICATIONS
# ═══════════════════════════════════════════════════════════════════════════
ws11 = wb.create_sheet("Applications"); ws11.sheet_properties.tabColor = C_RED
h11=["Company","Role","Level","CTC Target","Status","Applied","Referral?","Recruiter","OA Date","OA Result","R1 Date","R1 Type","R1 Result","R2 Date","R2 Type","R2 Result","R3 Date","R3 Type","R3 Result","Offer?","CTC Offered","Notes"]
w11=[16,20,8,12,13,11,9,18,11,11,11,13,11,11,13,11,11,13,11,9,12,28]
widths(ws11,w11)
for i,h in enumerate(h11,1): ws11.cell(1,i,value=h)
hdr(ws11,1,len(h11)); ws11.freeze_panes="A2"
cos=["Amazon","Microsoft","Google","Uber","Airbnb","Atlassian","Stripe","Databricks","Rubrik","Confluent","Razorpay","Flipkart","PhonePe","Zerodha","Groww","Swiggy"]
for idx,co in enumerate(cos):
    r=idx+2
    ws11.cell(r,1,value=co); ws11.cell(r,2,value="Senior SWE"); ws11.cell(r,3,value="SDE-2"); ws11.cell(r,4,value="40+ LPA"); ws11.cell(r,5,value="Not Applied")
    row_style(ws11,r,len(h11),idx%2==0)
dv(ws11,"E",2,20,["Not Applied","Applied","OA Scheduled","Interviewing","Offer","Rejected","Withdrew"])

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 12: REVISION TRACKER
# ═══════════════════════════════════════════════════════════════════════════
ws12 = wb.create_sheet("Revision Tracker"); ws12.sheet_properties.tabColor = C_YELLOW
h12=["#","Problem","Pattern","Date Solved","Day1","Day3","Day7","Day14","Day30","Forgotten","Mastered?"]
w12=[5,36,17,13,10,10,10,10,10,12,10]
widths(ws12,w12)
for i,h in enumerate(h12,1): ws12.cell(1,i,value=h)
hdr(ws12,1,len(h12)); ws12.freeze_panes="A2"
for idx,(num,name,lc,diff,pat,freq,ph) in enumerate(PROBLEMS):
    r=idx+2
    for ci,v in enumerate([num,name,pat,"","","","","","",0,""],1): ws12.cell(r,ci,value=v)
    row_style(ws12,r,len(h12),idx%2==0)
dv(ws12,"K",2,len(PROBLEMS)+1,["Yes","No"])

# ═══════════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════════
fp="/Users/sbaskar/Personal/Interview-Preparation-Tracker.xlsx"
wb.save(fp)
print(f"Created: {fp}")
print(f"Sheets: {wb.sheetnames}")
print(f"DSA Problems: {len(PROBLEMS)} | Java Topics: {len(java_topics)} | SD Topics: {len(sd_topics)}")
print(f"Kafka/Redis: {len(kr_topics)} | Cloud: {len(cloud_topics)} | Books: {len(books)} | Blogs: {len(eng_blogs)}")
